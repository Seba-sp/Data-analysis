"""Unified assessment mapper: assessment_id -> (report_type, assessment_type)."""

import logging
import os
import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv()
logger = logging.getLogger(__name__)

# Env var names for diagnosticos assessment IDs
_DIAG_MAPPING: Dict[str, Tuple[str, str]] = {
    "M1_ASSESSMENT_ID": ("diagnosticos", "M1"),
    "CL_ASSESSMENT_ID": ("diagnosticos", "CL"),
    "CIEN_ASSESSMENT_ID": ("diagnosticos", "CIEN"),
    "HYST_ASSESSMENT_ID": ("diagnosticos", "HYST"),
}

# Env var names for diagnosticos_uim assessment IDs
_UIM_MAPPING: Dict[str, Tuple[str, str]] = {
    "M1_UIM_ASSESSMENT_ID": ("diagnosticos_uim", "M1"),
    "F30M_ASSESSMENT_ID": ("diagnosticos_uim", "F30M"),
    "B30M_ASSESSMENT_ID": ("diagnosticos_uim", "B30M"),
    "Q30M_ASSESSMENT_ID": ("diagnosticos_uim", "Q30M"),
    "HYST_UIM_ASSESSMENT_ID": ("diagnosticos_uim", "HYST"),
}

_VALID_HEX_ID_RE = re.compile(r"^[a-fA-F0-9]{24}$")
# CL: included for ids.xlsx name-pattern parsing. CL assessment IDs are also routed via
# _DIAG_MAPPING env-var entries (CL_ASSESSMENT_ID) which take priority at lookup time.
# L30M: 3 rows exist in ids.xlsx as of 2026-03-08 but their IDs are invalid non-hex
# placeholders — all L30M rows are rejected at _register_route with "invalid_assessment_id"
# until valid hex IDs are provided. L30M is intentionally excluded from this set to keep
# rejection reason as "unsupported_group" so the startup summary log surfaces the issue.
_ALLOWED_GROUPS = {"M1", "M2", "H30M", "Q30M", "F30M", "B30M", "CL"}
_GROUP_ALIASES = {
    "M30M2": "M2",
    "M30M1": "M1",
    "M30M": "M1",
}
_TYPE_TO_REPORT = {
    "TEST DE EJE": "test_de_eje",
    "EXAMEN DE EJE": "examen_de_eje",
    "ENSAYO": "ensayo",
}
_PRODUCTION_MARKERS = {"production", "prod"}


class AssessmentMapper:
    """Maps hex assessment IDs to (report_type, assessment_type) tuples.

    Reads hex values from environment variables at construction time.
    Both diagnosticos and diagnosticos_uim mappings are merged into a single
    lookup dict keyed by hex assessment ID.
    """

    def __init__(self):
        """Build the routes dict from environment variables.

        Entries with a None env var value are silently skipped.
        If the same hex value appears in both DIAG and UIM mappings
        (i.e. the same hex is used by both systems), the UIM entry wins.
        """
        self._routes: Dict[str, Tuple[str, str]] = {}
        self._names: Dict[str, str] = {}       # normalized_id -> assessment_name
        self._rejected_names: List[str] = []   # names rejected for invalid_assessment_id
        self.validation_counters: Dict[str, int] = {"accepted": 0, "rejected": 0}
        self.validation_errors: Dict[str, int] = {}
        self.mapping_source = self._select_source()
        # Load diagnosticos routes first
        for env_var, route in _DIAG_MAPPING.items():
            hex_id = os.getenv(env_var)
            if hex_id is not None:
                self._routes[hex_id.lower()] = route

        # Load diagnosticos_uim routes (override on collision)
        for env_var, route in _UIM_MAPPING.items():
            hex_id = os.getenv(env_var)
            if hex_id is not None:
                self._routes[hex_id.lower()] = route

        self._load_ids_routes()

    def _is_production_environment(self) -> bool:
        """Infer production mode from common deployment environment markers."""
        for env_var in ("ENV", "APP_ENV", "ENVIRONMENT", "FLASK_ENV"):
            value = os.getenv(env_var, "").strip().lower()
            if value in _PRODUCTION_MARKERS:
                return True

        # Cloud Run service marker is present in production deployments.
        if os.getenv("K_SERVICE"):
            return True
        return False

    def _select_source(self) -> str:
        """Choose ids mapping source based on explicit override and environment."""
        explicit = os.getenv("ASSESSMENT_MAPPING_SOURCE", "").strip().lower()
        if explicit in {"local", "gcs"}:
            return explicit
        if self._is_production_environment():
            return "gcs"
        return "local"

    def _resolve_local_ids_path(self) -> Path:
        configured = os.getenv("IDS_XLSX_LOCAL_PATH", "inputs/ids.xlsx").strip()
        return Path(configured)

    def _resolve_gcs_target(self) -> Tuple[str, str]:
        configured = os.getenv("IDS_XLSX_GCS_PATH", "").strip()
        if not configured:
            raise ValueError("IDS_XLSX_GCS_PATH is required when source is gcs")

        if configured.startswith("gs://"):
            no_scheme = configured[5:]
            if "/" not in no_scheme:
                raise ValueError(
                    "IDS_XLSX_GCS_PATH must include object path (gs://bucket/object)"
                )
            bucket_name, object_path = no_scheme.split("/", 1)
            return bucket_name, object_path

        bucket_name = os.getenv("GCP_BUCKET_NAME", "").strip()
        if not bucket_name:
            raise ValueError(
                "GCP_BUCKET_NAME is required for non-gs:// IDS_XLSX_GCS_PATH values"
            )
        return bucket_name, configured.replace("\\", "/")

    def _read_local_ids_xlsx_bytes(self) -> bytes:
        local_path = self._resolve_local_ids_path()
        if not local_path.exists():
            raise FileNotFoundError(f"Local ids.xlsx not found: {local_path}")
        return local_path.read_bytes()

    def _read_gcs_ids_xlsx_bytes(self) -> bytes:
        bucket_name, object_path = self._resolve_gcs_target()
        from google.cloud import storage

        client = storage.Client()
        blob = client.bucket(bucket_name).blob(object_path)
        if not blob.exists():
            raise FileNotFoundError(
                f"GCS ids.xlsx object not found: gs://{bucket_name}/{object_path}"
            )
        return blob.download_as_bytes()

    def _read_ids_xlsx_bytes(self) -> bytes:
        if self.mapping_source == "gcs":
            return self._read_gcs_ids_xlsx_bytes()
        return self._read_local_ids_xlsx_bytes()

    def _load_ids_xlsx_rows(self, workbook_bytes: bytes) -> List[Tuple[int, str, str]]:
        """Load rows from ids.xlsx accepting headered and headerless layouts."""
        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        first_row = rows[0]
        first_row_normalized = [
            str(cell).strip().lower() if cell is not None else "" for cell in first_row
        ]
        has_headers = (
            "assessment_name" in first_row_normalized
            and "assessment_id" in first_row_normalized
        )

        if has_headers:
            name_idx = first_row_normalized.index("assessment_name")
            id_idx = first_row_normalized.index("assessment_id")
            start_index = 2
        else:
            name_idx = 0
            id_idx = 1
            start_index = 1

        parsed_rows: List[Tuple[int, str, str]] = []
        for row_number, row in enumerate(rows[start_index - 1 :], start=start_index):
            if row is None:
                continue
            name_cell = row[name_idx] if len(row) > name_idx else None
            id_cell = row[id_idx] if len(row) > id_idx else None
            if name_cell is None and id_cell is None:
                continue

            assessment_name = str(name_cell).strip() if name_cell is not None else ""
            assessment_id = str(id_cell).strip() if id_cell is not None else ""
            parsed_rows.append((row_number, assessment_id, assessment_name))

        return parsed_rows

    def _load_ids_routes(self) -> None:
        """Merge ids.xlsx-derived routes with existing env routes."""
        try:
            workbook_bytes = self._read_ids_xlsx_bytes()
        except FileNotFoundError as exc:
            logger.warning(str(exc), extra={"mapping_source": self.mapping_source})
            return
        except Exception as exc:
            logger.error(
                "Failed to load ids.xlsx mapping source",
                extra={"mapping_source": self.mapping_source, "error": str(exc)},
            )
            return

        rows = self._load_ids_xlsx_rows(workbook_bytes)
        accepted_before = self.validation_counters["accepted"]
        rejected_before = self.validation_counters["rejected"]
        for row_index, assessment_id, assessment_name in rows:
            self.register_ids_row(assessment_id, assessment_name, row_index=row_index)

        accepted_new = self.validation_counters["accepted"] - accepted_before
        rejected_new = self.validation_counters["rejected"] - rejected_before
        logger.info(
            "Loaded ids.xlsx routes",
            extra={
                "mapping_source": self.mapping_source,
                "rows_total": len(rows),
                "accepted_new": accepted_new,
                "rejected_new": rejected_new,
            },
        )
        # Always emit a startup summary warning so Cloud Run logs show which assessment
        # names were rejected due to invalid_assessment_id. Emitted unconditionally
        # (even when rejected_names is empty) so log presence is predictable.
        logger.warning(
            "ids.xlsx startup summary",
            extra={
                "mapping_source": self.mapping_source,
                "accepted": accepted_new,
                "rejected": rejected_new,
                "rejected_names": list(self._rejected_names),
            },
        )

    def _normalize_text(self, value: str) -> str:
        """Normalize text for deterministic parser behavior."""
        if not value:
            return ""
        text = unicodedata.normalize("NFKD", value)
        text = "".join(char for char in text if not unicodedata.combining(char))
        text = text.upper()
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def _parse_assessment_name(self, name: str) -> Optional[Tuple[str, str, str]]:
        """Parse `[GROUP]-[TYPE N]-DATA` into `(group, report_type, assessment_type)`."""
        normalized = self._normalize_text(name)
        normalized = re.sub(r"\s*-\s*", "-", normalized)
        match = re.match(r"^([A-Z0-9]+)-(.+)-DATA$", normalized)
        if not match:
            logger.warning(
                "Rejected assessment row: invalid name pattern",
                extra={"reason": "invalid_pattern", "assessment_name": name},
            )
            return None

        raw_group = match.group(1)
        group = _GROUP_ALIASES.get(raw_group, raw_group)
        type_with_number = match.group(2).strip()
        if group not in _ALLOWED_GROUPS:
            logger.warning(
                "Rejected assessment row: unsupported group",
                extra={
                    "reason": "unsupported_group",
                    "assessment_name": name,
                    "group": raw_group,
                },
            )
            return None

        type_match = re.match(r"^(.+?)\s+(\d+)$", type_with_number)
        if not type_match:
            logger.warning(
                "Rejected assessment row: invalid type segment",
                extra={"reason": "invalid_type_segment", "assessment_name": name},
            )
            return None

        report_type = _TYPE_TO_REPORT.get(type_match.group(1).strip())
        if report_type is None:
            logger.warning(
                "Rejected assessment row: unsupported type",
                extra={"reason": "unsupported_type", "assessment_name": name},
            )
            return None

        assessment_type = group
        return (group, report_type, assessment_type)

    def _record_rejection(
        self,
        reason: str,
        row_index: Optional[int],
        assessment_id: Optional[str],
        assessment_name: Optional[str],
    ) -> None:
        self.validation_counters["rejected"] += 1
        self.validation_errors[reason] = self.validation_errors.get(reason, 0) + 1
        # Track names rejected specifically for invalid_assessment_id so the startup
        # summary can surface which assessments are missing valid hex IDs.
        if reason == "invalid_assessment_id" and assessment_name:
            self._rejected_names.append(assessment_name)
        logger.warning(
            "Rejected ids mapping row",
            extra={
                "reason": reason,
                "row_index": row_index,
                "assessment_id": assessment_id,
                "assessment_name": assessment_name,
            },
        )

    def _register_route(
        self,
        assessment_id: Optional[str],
        report_type: str,
        assessment_type: str,
        row_index: Optional[int] = None,
        assessment_name: Optional[str] = None,
    ) -> bool:
        """Register one route with row-level validation and conflict safety."""
        if not assessment_id:
            self._record_rejection(
                "missing_assessment_id", row_index, assessment_id, assessment_name
            )
            return False

        normalized_id = assessment_id.strip().lower()
        if not _VALID_HEX_ID_RE.match(normalized_id):
            self._record_rejection(
                "invalid_assessment_id", row_index, assessment_id, assessment_name
            )
            return False

        candidate = (report_type, assessment_type)
        existing = self._routes.get(normalized_id)
        if existing is not None and existing != candidate:
            self._record_rejection(
                "conflicting_duplicate_id", row_index, normalized_id, assessment_name
            )
            return False

        if existing is None:
            self._routes[normalized_id] = candidate
            if assessment_name:
                self._names[normalized_id] = assessment_name

        self.validation_counters["accepted"] += 1
        return True

    def register_ids_row(
        self,
        assessment_id: Optional[str],
        assessment_name: Optional[str],
        row_index: Optional[int] = None,
    ) -> bool:
        """Validate and register one ids.xlsx row."""
        if not assessment_name or not assessment_name.strip():
            self._record_rejection(
                "missing_assessment_name", row_index, assessment_id, assessment_name
            )
            return False

        parsed = self._parse_assessment_name(assessment_name)
        if parsed is None:
            self._record_rejection(
                "invalid_assessment_name", row_index, assessment_id, assessment_name
            )
            return False

        _, report_type, assessment_type = parsed
        return self._register_route(
            assessment_id=assessment_id,
            report_type=report_type,
            assessment_type=assessment_type,
            row_index=row_index,
            assessment_name=assessment_name,
        )

    def get_route(self, assessment_id: str) -> Optional[Tuple[str, str]]:
        """Return (report_type, assessment_type) for a known assessment ID.

        Args:
            assessment_id: 24-character hex assessment ID from LearnWorlds.

        Returns:
            (report_type, assessment_type) tuple, or None if unknown.
        """
        if not assessment_id:
            return None
        return self._routes.get(assessment_id.lower())

    def get_route_full(self, assessment_id: str) -> Optional[Tuple[str, str, str]]:
        """Return (report_type, assessment_type, assessment_name) for a known assessment ID.

        Extends get_route() with the assessment_name so webhook handlers can tag
        queued students with a human-readable label without a second lookup.

        Args:
            assessment_id: 24-character hex assessment ID from LearnWorlds.

        Returns:
            (report_type, assessment_type, assessment_name) 3-tuple, or None if unknown.
            assessment_name is empty string if the name was not stored at registration time.
        """
        if not assessment_id:
            return None
        route = self._routes.get(assessment_id.lower())
        if route is None:
            return None
        name = self._names.get(assessment_id.lower(), "")
        return (route[0], route[1], name)

    def extract_assessment_id(self, url: str) -> Optional[str]:
        """Extract assessment ID from a LearnWorlds URL.

        Pattern: unit=<24-character hex id>

        Args:
            url: LearnWorlds assessment URL.

        Returns:
            24-character hex assessment ID, or None if not found.
        """
        if not url:
            return None
        match = re.search(r"unit=([a-fA-F0-9]{24})", url)
        if match:
            return match.group(1).lower()
        return None

    def is_valid_assessment_id(self, assessment_id: str) -> bool:
        """Check whether assessment_id is in the known routes.

        Args:
            assessment_id: Hex assessment ID to validate.

        Returns:
            True if the ID maps to a known route.
        """
        if not assessment_id:
            return False
        return assessment_id.lower() in self._routes
