"""Test de Habilidad plugin.

Builds a per-student reading-task status report from:
1) LearnWorlds responses (via ids.xlsx test-de-habilidad mappings), and
2) XLSX question banks in inputs/ with columns:
   [pregunta, alternativa, tarea_lectora, habilidad]

State logic:
  correct / total >= 80% -> "Dominada"
  otherwise              -> "En desarrollo"

The rendered table appends two fixed rows at the end:
  "Guias tematicas" | "Realizar"
  "Examen"          | "Realizar"
"""

from __future__ import annotations

import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from core.storage import StorageClient
from openpyxl import load_workbook
from weasyprint import HTML

from core.assessment_downloader import AssessmentDownloader
from core.assessment_mapper import AssessmentMapper
from reports.base import BaseReportGenerator
from reports.template_contracts import load_body_template
from reports.template_renderer import insert_dynamic_tables, render_with_placeholders

logger = logging.getLogger(__name__)

TASK_MASTERY_PERCENT = 80.0
ESTADO_DOMINADA = "Dominada"
ESTADO_EN_DESARROLLO = "En desarrollo"
ESTADO_REALIZAR = "Realizar"

IDS_LOCAL_PATH = Path("inputs/ids.xlsx")
BANKS_DIR = Path("inputs")
REPORT_TYPE = "test_de_habilidad"

_VALID_HEX_ID_RE = re.compile(r"^[a-fA-F0-9]{24}$")
_TDH_NAME_RE = re.compile(r"^([A-Z0-9]+)-TEST DE HABILIDAD\s+(\d+)-DATA$")
_BODY_RE = re.compile(r"<body[^>]*>(?P<body>[\s\S]*?)</body>", re.IGNORECASE)
_STYLE_RE = re.compile(r"<style[^>]*>(?P<style>[\s\S]*?)</style>", re.IGNORECASE)


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip()


def _display_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return unicodedata.normalize("NFC", text).strip()


def _normalize_answer(value: Any) -> str:
    return _normalize_text(value).upper()


def _safe_filename_component(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", value.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(".")
    return cleaned or "unknown"


def _extract_head_styles(html: str) -> str:
    return "\n".join(match.group("style") for match in _STYLE_RE.finditer(html))


def _extract_body_inner(html: str) -> str:
    match = _BODY_RE.search(html)
    if not match:
        raise ValueError("Template has no <body> section")
    return match.group("body")


def _assign_estado(percent: float) -> str:
    return ESTADO_DOMINADA if percent >= TASK_MASTERY_PERCENT else ESTADO_EN_DESARROLLO


def _lookup_answer_column(row: pd.Series, question_label: str) -> str:
    wanted = _normalize_text(question_label).lower()
    for col in row.index:
        if _normalize_text(col).lower() == wanted:
            return str(row[col]) if col in row else ""
    return ""


def _compose_cover_plus_body_html(cover_html: str, body_html: str) -> str:
    cover_styles = _extract_head_styles(cover_html)
    body_styles = _extract_head_styles(body_html)
    cover_body = _extract_body_inner(cover_html)
    body_inner = _extract_body_inner(body_html)

    return (
        "<!DOCTYPE html>"
        "<html lang=\"es\">"
        "<head>"
        "<meta charset=\"utf-8\" />"
        "<title>Test de Habilidad</title>"
        f"<style>{cover_styles}\n{body_styles}</style>"
        "</head>"
        "<body>"
        f"{cover_body}"
        "<div style=\"page-break-after: always;\"></div>"
        f"{body_inner}"
        "</body>"
        "</html>"
    )


@dataclass
class TareaStats:
    name: str
    total: int = 0
    correct: int = 0

    @property
    def percent(self) -> float:
        if self.total == 0:
            return 0.0
        return (self.correct / self.total) * 100.0


@dataclass
class HabilidadPlan:
    assessment_type: str
    student_id: str
    email: str
    assessment_name: str = ""
    habilidad_name: str = ""
    tareas: dict[str, TareaStats] = field(default_factory=dict)
    tarea_order: list[str] = field(default_factory=list)


@dataclass
class MappingRow:
    assessment_name: str
    assessment_type: str
    assessment_number: int
    assessment_id: str
    bank_path: Path


class TestDeHabilidadGenerator(BaseReportGenerator):
    def __init__(self) -> None:
        super().__init__(REPORT_TYPE)
        self.downloader = AssessmentDownloader(data_dir=f"data/{REPORT_TYPE}")
        self.templates_path = Path("templates") / REPORT_TYPE
        self.mapper = AssessmentMapper()
        self._available_bank_names_cache: Optional[set[str]] = None

    def _gcs_bank_blob(self, bank_name: str):
        from google.cloud import storage
        bucket_name, _ = self.mapper._resolve_gcs_target()
        prefix = os.getenv("BANKS_GCS_PREFIX", "inputs/").rstrip("/") + "/"
        return storage.Client().bucket(bucket_name).blob(f"{prefix}{bank_name}")

    def _available_bank_names(self) -> set[str]:
        if self._available_bank_names_cache is not None:
            return self._available_bank_names_cache

        if self.mapper.mapping_source == "gcs":
            from google.cloud import storage
            bucket_name, _ = self.mapper._resolve_gcs_target()
            prefix = os.getenv("BANKS_GCS_PREFIX", "inputs/").rstrip("/") + "/"
            blobs = storage.Client().bucket(bucket_name).list_blobs(prefix=prefix)
            names = {Path(blob.name).name for blob in blobs if blob.name and not blob.name.endswith("/")}
        else:
            names = {p.name for p in BANKS_DIR.glob("*.xlsx") if p.is_file()}

        self._available_bank_names_cache = names
        return names

    def _read_bank_bytes(self, bank_name: str, local_path: Optional[Path] = None) -> bytes:
        if self.mapper.mapping_source == "gcs":
            return self._gcs_bank_blob(bank_name).download_as_bytes()
        path = local_path if local_path is not None else BANKS_DIR / bank_name
        return path.read_bytes()

    def _load_test_de_habilidad_mapping(self) -> list[MappingRow]:
        if self.mapper.mapping_source == "local" and IDS_LOCAL_PATH.exists():
            ids_bytes = IDS_LOCAL_PATH.read_bytes()
            workbook = load_workbook(filename=BytesIO(ids_bytes), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        else:
            rows = [
                (assessment_name, assessment_id)
                for _, assessment_id, assessment_name in self.mapper.get_ids_rows()
            ]
        if not rows:
            return []

        first_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        has_headers = "assessment_name" in first_row and "assessment_id" in first_row
        if has_headers:
            name_idx = first_row.index("assessment_name")
            id_idx = first_row.index("assessment_id")
            data_rows = rows[1:]
        else:
            name_idx = 0
            id_idx = 1
            data_rows = rows

        mapping: list[MappingRow] = []
        for row in data_rows:
            if row is None:
                continue
            raw_name = row[name_idx] if len(row) > name_idx else None
            raw_id = row[id_idx] if len(row) > id_idx else None
            if not raw_name or not raw_id:
                continue

            normalized_name = _normalize_text(raw_name).upper()
            normalized_name = re.sub(r"\s*-\s*", "-", normalized_name)
            parsed = _TDH_NAME_RE.match(normalized_name)
            if not parsed:
                continue

            assessment_type = parsed.group(1)
            assessment_number = int(parsed.group(2))
            assessment_id = str(raw_id).strip().lower()
            if not _VALID_HEX_ID_RE.match(assessment_id):
                continue

            bank_name = f"{assessment_type}-TEST DE HABILIDAD {assessment_number}-DATA.xlsx"
            bank_path = BANKS_DIR / bank_name
            if bank_name not in self._available_bank_names():
                logger.warning("Question bank not found for mapping row: %s", bank_name)
                continue

            mapping.append(
                MappingRow(
                    assessment_name=normalized_name,
                    assessment_type=assessment_type,
                    assessment_number=assessment_number,
                    assessment_id=assessment_id,
                    bank_path=bank_path,
                )
            )

        mapping.sort(key=lambda item: (item.assessment_type, item.assessment_number))
        if not mapping:
            raise ValueError("No valid TEST DE HABILIDAD rows found in ids.xlsx")
        return mapping

    def download(self, assessment_name: str = "") -> dict[str, pd.DataFrame]:
        mapping = self._load_test_de_habilidad_mapping()
        if assessment_name:
            normalized_filter = _normalize_text(assessment_name).upper()
            normalized_filter = re.sub(r"\s*-\s*", "-", normalized_filter)
            mapping = [r for r in mapping if r.assessment_name == normalized_filter]
            if not mapping:
                logger.warning(
                    "No mapping rows found for assessment_name=%r — returning empty download",
                    assessment_name,
                )
        downloaded: dict[str, pd.DataFrame] = {}
        for item in mapping:
            key = f"{item.assessment_type}_TEST_DE_HABILIDAD_{item.assessment_number}"
            result = self.downloader.download_and_process_assessment(
                assessment_id=item.assessment_id,
                assessment_name=key,
                incremental_mode=False,
            )
            if isinstance(result, pd.DataFrame):
                df = result
            else:
                csv_path = self.downloader.get_csv_file_path(key)
                df = StorageClient().read_csv(csv_path, sep=";")
            downloaded[key] = df
        return downloaded

    def analyze(self, download_result: dict[str, pd.DataFrame]) -> dict[tuple[str, str], HabilidadPlan]:
        mapping = self._load_test_de_habilidad_mapping()
        mapping_by_key = {
            f"{m.assessment_type}_TEST_DE_HABILIDAD_{m.assessment_number}": m for m in mapping
        }

        student_plans: dict[tuple[str, str], HabilidadPlan] = {}

        for key, df in download_result.items():
            map_row = mapping_by_key.get(key)
            if map_row is None:
                continue

            bank_df = pd.read_excel(
                BytesIO(self._read_bank_bytes(map_row.bank_path.name, local_path=map_row.bank_path))
            )
            bank_df.columns = [_normalize_text(c).lower() for c in bank_df.columns]
            required_cols = {"pregunta", "alternativa", "tarea_lectora"}
            if not required_cols.issubset(set(bank_df.columns)):
                missing = required_cols - set(bank_df.columns)
                raise ValueError(f"Bank {map_row.bank_path.name} missing columns: {sorted(missing)}")

            # Derive habilidad name from bank if column present, else from assessment name
            habilidad_name = ""
            if "habilidad" in bank_df.columns:
                first_val = bank_df["habilidad"].dropna().iloc[0] if not bank_df["habilidad"].dropna().empty else ""
                habilidad_name = _display_text(first_val)

            for _, student_row in df.iterrows():
                email = _normalize_text(student_row.get("email") or student_row.get("username") or "")
                student_id = _normalize_text(student_row.get("user_id") or email)
                if not email:
                    continue

                plan_key = (map_row.assessment_type, email)
                if plan_key not in student_plans:
                    student_plans[plan_key] = HabilidadPlan(
                        assessment_type=map_row.assessment_type,
                        student_id=student_id,
                        email=email,
                        assessment_name=map_row.assessment_name,
                        habilidad_name=habilidad_name,
                    )
                plan = student_plans[plan_key]

                for _, question in bank_df.iterrows():
                    tarea_name = _display_text(question["tarea_lectora"])
                    question_label = _display_text(question["pregunta"])
                    correct_answer = _normalize_answer(question["alternativa"])
                    user_answer = _normalize_answer(_lookup_answer_column(student_row, question_label))
                    is_correct = user_answer == correct_answer

                    if tarea_name not in plan.tareas:
                        plan.tareas[tarea_name] = TareaStats(name=tarea_name)
                        plan.tarea_order.append(tarea_name)
                    tarea = plan.tareas[tarea_name]
                    tarea.total += 1
                    tarea.correct += 1 if is_correct else 0

        return student_plans

    def render(self, analysis_result: dict[tuple[str, str], HabilidadPlan]) -> Path:
        output_dir = self.data_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        cover_html = (self.templates_path / "cover.html").read_text(encoding="utf-8")
        body_template = load_body_template(REPORT_TYPE)

        for (_, email), plan in analysis_result.items():
            ordered_tareas = [plan.tareas[t] for t in plan.tarea_order if t in plan.tareas]
            if not ordered_tareas:
                continue

            # Dynamic rows + two fixed rows appended at the end
            tarea_rows = [
                {
                    "tarea_lectora": t.name,
                    "estado": _assign_estado(t.percent),
                }
                for t in ordered_tareas
            ]
            tarea_rows.append({"tarea_lectora": "Gu\u00edas tem\u00e1ticas", "estado": ESTADO_REALIZAR})
            tarea_rows.append({"tarea_lectora": "Examen", "estado": ESTADO_REALIZAR})

            computed_values = {
                "habilidad_name": plan.habilidad_name or plan.assessment_type,
            }
            rendered_body = render_with_placeholders(
                report_type=REPORT_TYPE,
                body_html=body_template,
                computed_values=computed_values,
                static_values={},
            )
            rendered_body = insert_dynamic_tables(
                REPORT_TYPE,
                rendered_body,
                {"tarea_status_rows": tarea_rows},
            )
            final_html = _compose_cover_plus_body_html(cover_html, rendered_body)

            pdf_bytes = HTML(string=final_html, base_url=str(Path.cwd())).write_pdf()
            assessment_label = plan.assessment_name or plan.assessment_type
            pdf_path = output_dir / (
                f"informe_{_safe_filename_component(REPORT_TYPE)}"
                f"_{_safe_filename_component(assessment_label)}"
                f"_{_safe_filename_component(email)}.pdf"
            )
            pdf_path.write_bytes(pdf_bytes)

        return output_dir
