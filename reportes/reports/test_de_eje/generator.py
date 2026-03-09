"""Test de Eje plugin.

Builds a per-student study plan from:
1) LearnWorlds responses (via ids.xlsx test-de-eje mappings), and
2) XLSX question banks in inputs/ with columns:
   [pregunta, alternativa, unidad, leccion]
"""

from __future__ import annotations

import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from core.storage import StorageClient
from openpyxl import load_workbook
from weasyprint import HTML

from core.assessment_downloader import AssessmentDownloader
from core.assessment_mapper import AssessmentMapper
from reports.base import BaseReportGenerator
from reports.template_contracts import load_body_template
from reports.template_renderer import render_with_placeholders

logger = logging.getLogger(__name__)


# Thresholds (kept as variables for easy future tuning)
LECTURE_PASS_PERCENT = 70.0
UNIT_FULL_MASTERY_PERCENT = 100.0

# Duration model (hours)
HOURS_PER_LECTURE = 1.0
HOURS_PER_UNIT_TEST = 0.5
HOURS_PER_TOPIC_GUIDE = 2.0
HOURS_PER_EXAM = 2.0

TICK = "\u2713"
SQUARE = "\u25A1"
NOT_REQUIRED = "No requerido"
MAX_ACTIVITY_ROWS_PER_PAGE = 14
UNIT_CONTENT_WIDTH_MM = 160
UNIT_ACTIVITY_COL_MM = 96
UNIT_ACTION_COL_MM = 64

IDS_LOCAL_PATH = Path("inputs/ids.xlsx")
BANKS_DIR = Path("inputs")
REPORT_TYPE = "test_de_eje"

_VALID_HEX_ID_RE = re.compile(r"^[a-fA-F0-9]{24}$")
_TDE_NAME_RE = re.compile(r"^([A-Z0-9]+)-TEST DE EJE\s+(\d+)-DATA$")
_BODY_RE = re.compile(r"<body[^>]*>(?P<body>[\s\S]*?)</body>", re.IGNORECASE)
_STYLE_RE = re.compile(r"<style[^>]*>(?P<style>[\s\S]*?)</style>", re.IGNORECASE)
_UNIT_SECTION_RE = re.compile(
    r"<section class=\"page unit\"[\s\S]*?</section>",
    re.IGNORECASE,
)


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip()


def _normalize_answer(value: Any) -> str:
    return _normalize_text(value).upper()


def _safe_filename_component(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", value.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(".")
    return cleaned or "unknown"


def _strip_data_suffix(value: str) -> str:
    return re.sub(r"-DATA$", "", value.strip(), flags=re.IGNORECASE)


def _extract_head_styles(html: str) -> str:
    return "\n".join(match.group("style") for match in _STYLE_RE.finditer(html))


def _extract_body_inner(html: str) -> str:
    match = _BODY_RE.search(html)
    if not match:
        raise ValueError("Template has no <body> section")
    return match.group("body")


@dataclass
class LessonStats:
    total: int = 0
    correct: int = 0

    @property
    def percent(self) -> float:
        if self.total == 0:
            return 0.0
        return (self.correct / self.total) * 100.0


@dataclass
class UnitProgress:
    name: str
    total: int = 0
    correct: int = 0
    lessons: dict[str, LessonStats] = field(default_factory=dict)

    @property
    def percent(self) -> float:
        if self.total == 0:
            return 0.0
        return (self.correct / self.total) * 100.0


@dataclass
class StudentPlan:
    assessment_type: str
    student_id: str
    email: str
    assessment_name: str = ""
    units: dict[str, UnitProgress] = field(default_factory=dict)


@dataclass
class MappingRow:
    assessment_name: str
    assessment_type: str
    assessment_number: int
    assessment_id: str
    bank_path: Path


def _lookup_answer_column(row: pd.Series, question_label: str) -> str:
    wanted = _normalize_text(question_label).lower()
    for col in row.index:
        if _normalize_text(col).lower() == wanted:
            return str(row[col]) if col in row else ""
    return ""


def _should_mark_completed(percent: float, threshold: float) -> bool:
    return percent >= threshold


def _is_unit_fully_mastered(unit_percent: float, full_mastery_threshold: float) -> bool:
    return unit_percent >= full_mastery_threshold


def _build_unit_activity_rows(unit: UnitProgress) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for lesson_name, lesson_stats in unit.lessons.items():
        lesson_done = _should_mark_completed(lesson_stats.percent, LECTURE_PASS_PERCENT)
        rows.append(
            {
                "activity": lesson_name,
                "action": TICK if lesson_done else SQUARE,
            }
        )

    fully_mastered = _is_unit_fully_mastered(unit.percent, UNIT_FULL_MASTERY_PERCENT)
    rows.append(
        {
            "activity": "Test de unidad",
            "action": NOT_REQUIRED if fully_mastered else SQUARE,
        }
    )
    rows.append(
        {
            "activity": "Guia tematica",
            "action": NOT_REQUIRED if fully_mastered else SQUARE,
        }
    )
    return rows


def _estimate_total_hours(units: list[UnitProgress]) -> float:
    pending_lectures = 0
    pending_unit_tests = 0
    pending_guides = 0
    for unit in units:
        for lesson_stats in unit.lessons.values():
            if not _should_mark_completed(lesson_stats.percent, LECTURE_PASS_PERCENT):
                pending_lectures += 1
        if not _is_unit_fully_mastered(unit.percent, UNIT_FULL_MASTERY_PERCENT):
            pending_unit_tests += 1
            pending_guides += 1

    return (
        pending_lectures * HOURS_PER_LECTURE
        + pending_unit_tests * HOURS_PER_UNIT_TEST
        + pending_guides * HOURS_PER_TOPIC_GUIDE
        + HOURS_PER_EXAM
    )


def _hours_to_str(hours: float) -> str:
    if float(hours).is_integer():
        return str(int(hours))
    return f"{hours:.1f}".rstrip("0").rstrip(".")


def _render_activity_table(rows: list[dict[str, str]]) -> str:
    body_rows = []
    for row in rows:
        body_rows.append(
            "<tr>"
            f"<td>{row['activity']}</td>"
            f"<td>{row['action']}</td>"
            "</tr>"
        )
    return (
        "<table class=\"activity-table\" aria-label=\"Actividades\" "
        f"style=\"display:table; width:{UNIT_CONTENT_WIDTH_MM}mm; min-width:{UNIT_CONTENT_WIDTH_MM}mm; "
        f"max-width:{UNIT_CONTENT_WIDTH_MM}mm; margin:0 0 8pt; table-layout:fixed;\">"
        f"<colgroup><col style=\"width:{UNIT_ACTIVITY_COL_MM}mm\" /><col style=\"width:{UNIT_ACTION_COL_MM}mm\" /></colgroup>"
        "<thead><tr><th>Actividad</th><th>Acci&oacute;n</th></tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table>"
    )


def _build_unit_sections(unit_rows: list[tuple[str, str, list[dict[str, str]]]]) -> str:
    # Dynamic pagination: pack multiple small units into the same page.
    pages: list[list[tuple[str, str, list[dict[str, str]]]]] = []
    current_page: list[tuple[str, str, list[dict[str, str]]]] = []
    current_rows = 0

    for unit_name, pd_text, rows in unit_rows:
        # row budget: table rows + 2 lines (title/domain) + spacing
        unit_budget = len(rows) + 3
        if current_page and (current_rows + unit_budget > MAX_ACTIVITY_ROWS_PER_PAGE):
            pages.append(current_page)
            current_page = []
            current_rows = 0

        current_page.append((unit_name, pd_text, rows))
        current_rows += unit_budget

    if current_page:
        pages.append(current_page)

    sections: list[str] = []
    for idx, page_units in enumerate(pages, start=2):
        cards: list[str] = []
        for unit_name, pd_text, rows in page_units:
            table_html = _render_activity_table(rows)
            cards.append(
                (
                    f"<div class=\"unit-card\" style=\"display:block; width:{UNIT_CONTENT_WIDTH_MM}mm; "
                    "margin:0 0 12pt;\">"
                    f"<p class=\"unit-title\">Unidad: <span>{unit_name}</span></p>"
                    f"<p class=\"unit-domain\">Dominio inicial: <span>{pd_text}</span></p>"
                    f"<div class=\"unit-table\" style=\"display:block; width:{UNIT_CONTENT_WIDTH_MM}mm; "
                    f"margin:0 0 8pt;\">{table_html}</div>"
                    "</div>"
                )
            )
        sections.append(
            (
                f"<section class=\"page unit\" data-page=\"{idx}\" data-placeholder=\"unit_block\">"
                "<div class=\"content\">"
                f"{''.join(cards)}"
                "</div>"
                "</section>"
            )
        )
    return "".join(sections)


def _replace_unit_sections(rendered_body_html: str, unit_rows: list[tuple[str, str, list[dict[str, str]]]]) -> str:
    matches = list(_UNIT_SECTION_RE.finditer(rendered_body_html))
    if not matches:
        raise ValueError("Could not locate unit sections in test_de_eje template")

    generated_sections = _build_unit_sections(unit_rows)
    start = matches[0].start()
    end = matches[-1].end()
    return rendered_body_html[:start] + generated_sections + rendered_body_html[end:]


def _compose_cover_plus_body_html(cover_html: str, body_html: str) -> str:
    cover_styles = _extract_head_styles(cover_html)
    body_styles = _extract_head_styles(body_html)
    cover_body = _extract_body_inner(cover_html)
    body_inner = _extract_body_inner(body_html)

    return (
        "<!DOCTYPE html>"
        "<html lang=\"es\">"
        "<head>"
        "<meta charset=\"utf-8\" />"
        "<title>Test de eje</title>"
        f"<style>{cover_styles}\n{body_styles}</style>"
        "</head>"
        "<body>"
        f"{cover_body}"
        "<div style=\"page-break-after: always;\"></div>"
        f"{body_inner}"
        "</body>"
        "</html>"
    )


class TestDeEjeGenerator(BaseReportGenerator):
    def __init__(self) -> None:
        super().__init__(REPORT_TYPE)
        self.downloader = AssessmentDownloader(data_dir=f"data/{REPORT_TYPE}")
        self.templates_path = Path("templates") / REPORT_TYPE
        self.mapper = AssessmentMapper()
        self._available_bank_names_cache: Optional[set[str]] = None

    def _gcs_bank_blob(self, bank_name: str):
        """Return a GCS Blob for the given question bank filename."""
        from google.cloud import storage
        bucket_name, _ = self.mapper._resolve_gcs_target()
        prefix = os.getenv("BANKS_GCS_PREFIX", "inputs/").rstrip("/") + "/"
        return storage.Client().bucket(bucket_name).blob(f"{prefix}{bank_name}")

    def _available_bank_names(self) -> set[str]:
        """Build a cached set of available bank filenames."""
        if self._available_bank_names_cache is not None:
            return self._available_bank_names_cache

        if self.mapper.mapping_source == "gcs":
            from google.cloud import storage

            bucket_name, _ = self.mapper._resolve_gcs_target()
            prefix = os.getenv("BANKS_GCS_PREFIX", "inputs/").rstrip("/") + "/"
            blobs = storage.Client().bucket(bucket_name).list_blobs(prefix=prefix)
            names = {Path(blob.name).name for blob in blobs if blob.name and not blob.name.endswith("/")}
        else:
            names = {p.name for p in BANKS_DIR.glob("*.xlsx") if p.is_file()}

        self._available_bank_names_cache = names
        return names

    def _read_bank_bytes(self, bank_name: str, local_path: Optional[Path] = None) -> bytes:
        """Download question bank bytes from GCS (production) or read from local path (dev).

        local_path, when provided, is used as-is in local mode (useful for tests that
        write banks to a temp directory rather than the default BANKS_DIR).
        """
        if self.mapper.mapping_source == "gcs":
            return self._gcs_bank_blob(bank_name).download_as_bytes()
        path = local_path if local_path is not None else BANKS_DIR / bank_name
        return path.read_bytes()

    def _load_test_de_eje_mapping(self) -> list[MappingRow]:
        if self.mapper.mapping_source == "local" and IDS_LOCAL_PATH.exists():
            ids_bytes = IDS_LOCAL_PATH.read_bytes()
            workbook = load_workbook(filename=BytesIO(ids_bytes), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        else:
            rows = [
                (assessment_name, assessment_id)
                for _, assessment_id, assessment_name in self.mapper.get_ids_rows()
            ]
        if not rows:
            return []

        first_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        has_headers = "assessment_name" in first_row and "assessment_id" in first_row
        if has_headers:
            name_idx = first_row.index("assessment_name")
            id_idx = first_row.index("assessment_id")
            data_rows = rows[1:]
        else:
            name_idx = 0
            id_idx = 1
            data_rows = rows

        mapping: list[MappingRow] = []
        for row in data_rows:
            if row is None:
                continue
            raw_name = row[name_idx] if len(row) > name_idx else None
            raw_id = row[id_idx] if len(row) > id_idx else None
            if not raw_name or not raw_id:
                continue

            normalized_name = _normalize_text(raw_name).upper()
            normalized_name = re.sub(r"\s*-\s*", "-", normalized_name)
            parsed = _TDE_NAME_RE.match(normalized_name)
            if not parsed:
                continue

            assessment_type = parsed.group(1)
            assessment_number = int(parsed.group(2))
            assessment_id = str(raw_id).strip().lower()
            if not _VALID_HEX_ID_RE.match(assessment_id):
                continue

            bank_name = f"{assessment_type}-TEST DE EJE {assessment_number}-DATA.xlsx"
            bank_path = BANKS_DIR / bank_name
            if bank_name not in self._available_bank_names():
                logger.warning("Question bank not found for mapping row: %s", bank_name)
                continue

            mapping.append(
                MappingRow(
                    assessment_name=normalized_name,
                    assessment_type=assessment_type,
                    assessment_number=assessment_number,
                    assessment_id=assessment_id,
                    bank_path=bank_path,
                )
            )

        mapping.sort(key=lambda item: (item.assessment_type, item.assessment_number))
        if not mapping:
            raise ValueError("No valid TEST DE EJE rows found in ids.xlsx")
        return mapping

    def download(self, assessment_name: str = "") -> dict[str, pd.DataFrame]:
        """Download assessment responses, optionally scoped to one assessment.

        Args:
            assessment_name: When non-empty, only download rows matching this
                normalized assessment_name from the ids.xlsx mapping.
                Empty string downloads all valid assessments.
        """
        mapping = self._load_test_de_eje_mapping()
        if assessment_name:
            normalized_filter = _normalize_text(assessment_name).upper()
            normalized_filter = re.sub(r"\s*-\s*", "-", normalized_filter)
            mapping = [r for r in mapping if r.assessment_name == normalized_filter]
            if not mapping:
                logger.warning(
                    "No mapping rows found for assessment_name=%r — returning empty download",
                    assessment_name,
                )
        downloaded: dict[str, pd.DataFrame] = {}
        for item in mapping:
            assessment_name = f"{item.assessment_type}_TEST_DE_EJE_{item.assessment_number}"
            result = self.downloader.download_and_process_assessment(
                assessment_id=item.assessment_id,
                assessment_name=assessment_name,
                incremental_mode=False,
            )
            if isinstance(result, pd.DataFrame):
                df = result
            else:
                csv_path = self.downloader.get_csv_file_path(assessment_name)
                df = StorageClient().read_csv(csv_path, sep=";")
            downloaded[assessment_name] = df
        return downloaded

    def analyze(self, download_result: dict[str, pd.DataFrame]) -> dict[tuple[str, str], StudentPlan]:
        mapping = self._load_test_de_eje_mapping()
        mapping_by_assessment_name = {
            f"{m.assessment_type}_TEST_DE_EJE_{m.assessment_number}": m for m in mapping
        }

        student_plans: dict[tuple[str, str], StudentPlan] = {}

        for assessment_name, df in download_result.items():
            map_row = mapping_by_assessment_name.get(assessment_name)
            if map_row is None:
                continue

            bank_df = pd.read_excel(BytesIO(self._read_bank_bytes(map_row.bank_path.name, local_path=map_row.bank_path)))
            bank_df.columns = [_normalize_text(c).lower() for c in bank_df.columns]
            required_cols = {"pregunta", "alternativa", "unidad", "leccion"}
            if not required_cols.issubset(set(bank_df.columns)):
                missing = required_cols - set(bank_df.columns)
                raise ValueError(f"Bank {map_row.bank_path.name} missing columns: {sorted(missing)}")

            for _, student_row in df.iterrows():
                email = _normalize_text(student_row.get("email") or student_row.get("username") or "")
                student_id = _normalize_text(student_row.get("user_id") or email)
                if not email:
                    continue

                plan_key = (map_row.assessment_type, email)
                if plan_key not in student_plans:
                    student_plans[plan_key] = StudentPlan(
                        assessment_type=map_row.assessment_type,
                        student_id=student_id,
                        email=email,
                        assessment_name=map_row.assessment_name,
                    )
                plan = student_plans[plan_key]

                for _, question in bank_df.iterrows():
                    unit_name = _normalize_text(question["unidad"])
                    lesson_name = _normalize_text(question["leccion"])
                    question_label = _normalize_text(question["pregunta"])
                    correct_answer = _normalize_answer(question["alternativa"])
                    user_answer = _normalize_answer(_lookup_answer_column(student_row, question_label))
                    is_correct = user_answer == correct_answer

                    if unit_name not in plan.units:
                        plan.units[unit_name] = UnitProgress(name=unit_name)
                    unit = plan.units[unit_name]
                    unit.total += 1
                    unit.correct += 1 if is_correct else 0

                    if lesson_name not in unit.lessons:
                        unit.lessons[lesson_name] = LessonStats()
                    lesson = unit.lessons[lesson_name]
                    lesson.total += 1
                    lesson.correct += 1 if is_correct else 0

        return student_plans

    def render(self, analysis_result: dict[tuple[str, str], StudentPlan]) -> Path:
        output_dir = self.data_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        cover_html = (self.templates_path / "cover.html").read_text(encoding="utf-8")
        body_template = load_body_template(REPORT_TYPE)

        for (_, email), plan in analysis_result.items():
            ordered_units = list(plan.units.values())
            if not ordered_units:
                continue

            unit_rows: list[tuple[str, str, list[dict[str, str]]]] = []
            for unit in ordered_units:
                unit_rows.append(
                    (
                        unit.name,
                        f"{unit.percent:.1f}%",
                        _build_unit_activity_rows(unit),
                    )
                )

            estimated_hours = _hours_to_str(_estimate_total_hours(ordered_units))

            computed_values = {
                "estimated_total_hours": estimated_hours,
                "unit_1_name": unit_rows[0][0] if len(unit_rows) > 0 else "-",
                "unit_1_initial_pd": unit_rows[0][1] if len(unit_rows) > 0 else "0%",
                "unit_1_activities_table": "",
                "unit_2_name": unit_rows[1][0] if len(unit_rows) > 1 else "-",
                "unit_2_initial_pd": unit_rows[1][1] if len(unit_rows) > 1 else "0%",
                "unit_2_activities_table": "",
                "unit_3_name": unit_rows[2][0] if len(unit_rows) > 2 else "-",
                "unit_3_initial_pd": unit_rows[2][1] if len(unit_rows) > 2 else "0%",
                "unit_3_activities_table": "",
            }
            static_values = {
                "report_title": "📍 Tu Plan de Estudio Personalizado",
                "cover_subtitle": "",
                "unit_block": "",
                "final_exam_heading": "",
            }
            rendered_body = render_with_placeholders(
                report_type=REPORT_TYPE,
                body_html=body_template,
                computed_values=computed_values,
                static_values=static_values,
            )
            rendered_body = _replace_unit_sections(rendered_body, unit_rows)
            final_html = _compose_cover_plus_body_html(cover_html, rendered_body)

            pdf_bytes = HTML(string=final_html, base_url=str(Path.cwd())).write_pdf()
            assessment_label = plan.assessment_name or plan.assessment_type
            pdf_path = output_dir / (
                f"informe_{_safe_filename_component(REPORT_TYPE)}"
                f"_{_safe_filename_component(assessment_label)}"
                f"_{_safe_filename_component(email)}.pdf"
            )
            pdf_path.write_bytes(pdf_bytes)

        return output_dir
