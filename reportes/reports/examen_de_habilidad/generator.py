"""Examen de Habilidad plugin.

Builds a per-student unit-level performance report from:
1) LearnWorlds responses (via ids.xlsx examen-de-habilidad mappings), and
2) XLSX question banks in inputs/ with columns:
   [pregunta, alternativa, unidad]
"""

from __future__ import annotations

import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from core.storage import StorageClient
from openpyxl import load_workbook
from weasyprint import HTML

from core.assessment_downloader import AssessmentDownloader
from core.assessment_mapper import AssessmentMapper
from reports.base import BaseReportGenerator
from reports.template_contracts import (
    discover_placeholders_in_html,
    load_body_template,
    load_report_placeholder_schema,
)
from reports.template_renderer import insert_dynamic_tables, render_with_placeholders

logger = logging.getLogger(__name__)

IDS_LOCAL_PATH = Path("inputs/ids.xlsx")
BANKS_DIR = Path("inputs")
REPORT_TYPE = "examen_de_habilidad"

_VALID_HEX_ID_RE = re.compile(r"^[a-fA-F0-9]{24}$")
_EDH_NAME_RE = re.compile(r"^([A-Z0-9]+)-EXAMEN DE HABILIDAD\s+(\d+)-DATA$")
_BODY_RE = re.compile(r"<body[^>]*>(?P<body>[\s\S]*?)</body>", re.IGNORECASE)
_STYLE_RE = re.compile(r"<style[^>]*>(?P<style>[\s\S]*?)</style>", re.IGNORECASE)


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip()


def _display_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return unicodedata.normalize("NFC", text).strip()


def _normalize_answer(value: Any) -> str:
    return _normalize_text(value).upper()


def _safe_filename_component(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", value.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(".")
    return cleaned or "unknown"


def _strip_data_suffix(value: str) -> str:
    return re.sub(r"-DATA$", "", value.strip(), flags=re.IGNORECASE)


def _extract_head_styles(html: str) -> str:
    return "\n".join(match.group("style") for match in _STYLE_RE.finditer(html))


def _extract_body_inner(html: str) -> str:
    match = _BODY_RE.search(html)
    if not match:
        raise ValueError("Template has no <body> section")
    return match.group("body")


def _assign_estado_recomendacion(pdu_percent: float) -> tuple[str, str]:
    if pdu_percent >= 80.0:
        return "Solido", "RS"
    if pdu_percent >= 50.0:
        return "En desarrollo", "RD"
    return "Riesgo", "RR"


def _compose_cover_plus_body_html(cover_html: str, body_html: str) -> str:
    cover_styles = _extract_head_styles(cover_html)
    body_styles = _extract_head_styles(body_html)
    cover_body = _extract_body_inner(cover_html)
    body_inner = _extract_body_inner(body_html)

    return (
        "<!DOCTYPE html>"
        "<html lang=\"es\">"
        "<head>"
        "<meta charset=\"utf-8\" />"
        "<title>Examen de Habilidad</title>"
        f"<style>{cover_styles}\n{body_styles}</style>"
        "</head>"
        "<body>"
        f"{cover_body}"
        "<div style=\"page-break-after: always;\"></div>"
        f"{body_inner}"
        "</body>"
        "</html>"
    )


@dataclass
class UnitStats:
    name: str
    total: int = 0
    correct: int = 0

    @property
    def percent(self) -> float:
        if self.total == 0:
            return 0.0
        return (self.correct / self.total) * 100.0


@dataclass
class ExamenPlan:
    assessment_type: str
    student_id: str
    email: str
    assessment_name: str = ""
    units: dict[str, UnitStats] = field(default_factory=dict)
    unit_order: list[str] = field(default_factory=list)


@dataclass
class MappingRow:
    assessment_name: str
    assessment_type: str
    assessment_number: int
    assessment_id: str
    bank_path: Path


def _lookup_answer_column(row: pd.Series, question_label: str) -> str:
    wanted = _normalize_text(question_label).lower()
    for col in row.index:
        if _normalize_text(col).lower() == wanted:
            return str(row[col]) if col in row else ""
    return ""


class ExamenDeHabilidadGenerator(BaseReportGenerator):
    def __init__(self) -> None:
        super().__init__(REPORT_TYPE)
        self.downloader = AssessmentDownloader(data_dir=f"data/{REPORT_TYPE}")
        self.templates_path = Path("templates") / REPORT_TYPE
        self.mapper = AssessmentMapper()
        self._available_bank_names_cache: Optional[set[str]] = None

    def _gcs_bank_blob(self, bank_name: str):
        """Return a GCS Blob for the given question bank filename."""
        from google.cloud import storage
        bucket_name, _ = self.mapper._resolve_gcs_target()
        prefix = os.getenv("BANKS_GCS_PREFIX", "inputs/").rstrip("/") + "/"
        return storage.Client().bucket(bucket_name).blob(f"{prefix}{bank_name}")

    def _available_bank_names(self) -> set[str]:
        """Build a cached set of available bank filenames."""
        if self._available_bank_names_cache is not None:
            return self._available_bank_names_cache

        if self.mapper.mapping_source == "gcs":
            from google.cloud import storage

            bucket_name, _ = self.mapper._resolve_gcs_target()
            prefix = os.getenv("BANKS_GCS_PREFIX", "inputs/").rstrip("/") + "/"
            blobs = storage.Client().bucket(bucket_name).list_blobs(prefix=prefix)
            names = {Path(blob.name).name for blob in blobs if blob.name and not blob.name.endswith("/")}
        else:
            names = {p.name for p in BANKS_DIR.glob("*.xlsx") if p.is_file()}

        self._available_bank_names_cache = names
        return names

    def _read_bank_bytes(self, bank_name: str, local_path: Optional[Path] = None) -> bytes:
        """Download question bank bytes from GCS (production) or read from local path (dev)."""
        if self.mapper.mapping_source == "gcs":
            return self._gcs_bank_blob(bank_name).download_as_bytes()
        path = local_path if local_path is not None else BANKS_DIR / bank_name
        return path.read_bytes()

    def _load_examen_de_habilidad_mapping(self) -> list[MappingRow]:
        if self.mapper.mapping_source == "local" and IDS_LOCAL_PATH.exists():
            ids_bytes = IDS_LOCAL_PATH.read_bytes()
            workbook = load_workbook(filename=BytesIO(ids_bytes), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        else:
            rows = [
                (assessment_name, assessment_id)
                for _, assessment_id, assessment_name in self.mapper.get_ids_rows()
            ]
        if not rows:
            raise ValueError("No valid EXAMEN DE HABILIDAD rows found in ids.xlsx")

        first_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        has_headers = "assessment_name" in first_row and "assessment_id" in first_row
        if has_headers:
            name_idx = first_row.index("assessment_name")
            id_idx = first_row.index("assessment_id")
            data_rows = rows[1:]
        else:
            name_idx = 0
            id_idx = 1
            data_rows = rows

        mapping: list[MappingRow] = []
        for row in data_rows:
            if row is None:
                continue
            raw_name = row[name_idx] if len(row) > name_idx else None
            raw_id = row[id_idx] if len(row) > id_idx else None
            if not raw_name or not raw_id:
                continue

            normalized_name = _normalize_text(raw_name).upper()
            normalized_name = re.sub(r"\s*-\s*", "-", normalized_name)
            parsed = _EDH_NAME_RE.match(normalized_name)
            if not parsed:
                continue

            assessment_type = parsed.group(1)
            assessment_number = int(parsed.group(2))
            assessment_id = str(raw_id).strip().lower()
            if not _VALID_HEX_ID_RE.match(assessment_id):
                continue

            bank_name = f"{assessment_type}-EXAMEN DE HABILIDAD {assessment_number}-DATA.xlsx"
            bank_path = BANKS_DIR / bank_name
            if bank_name not in self._available_bank_names():
                logger.warning("Question bank not found for mapping row: %s", bank_name)
                continue

            mapping.append(
                MappingRow(
                    assessment_name=normalized_name,
                    assessment_type=assessment_type,
                    assessment_number=assessment_number,
                    assessment_id=assessment_id,
                    bank_path=bank_path,
                )
            )

        mapping.sort(key=lambda item: (item.assessment_type, item.assessment_number))
        if not mapping:
            raise ValueError("No valid EXAMEN DE HABILIDAD rows found in ids.xlsx")
        return mapping

    def download(self, assessment_name: str = "") -> dict[str, pd.DataFrame]:
        """Download assessment responses, optionally scoped to one assessment."""
        mapping = self._load_examen_de_habilidad_mapping()
        if assessment_name:
            normalized_filter = _normalize_text(assessment_name).upper()
            normalized_filter = re.sub(r"\s*-\s*", "-", normalized_filter)
            mapping = [r for r in mapping if r.assessment_name == normalized_filter]
            if not mapping:
                logger.warning(
                    "No mapping rows found for assessment_name=%r — returning empty download",
                    assessment_name,
                )
        downloaded: dict[str, pd.DataFrame] = {}
        for item in mapping:
            assessment_name = f"{item.assessment_type}_EXAMEN_DE_HABILIDAD_{item.assessment_number}"
            result = self.downloader.download_and_process_assessment(
                assessment_id=item.assessment_id,
                assessment_name=assessment_name,
                incremental_mode=False,
            )
            if isinstance(result, pd.DataFrame):
                df = result
            else:
                csv_path = self.downloader.get_csv_file_path(assessment_name)
                df = StorageClient().read_csv(csv_path, sep=";")
            downloaded[assessment_name] = df
        return downloaded

    def analyze(self, download_result: dict[str, pd.DataFrame]) -> dict[tuple[str, str], ExamenPlan]:
        mapping = self._load_examen_de_habilidad_mapping()
        mapping_by_assessment_name = {
            f"{m.assessment_type}_EXAMEN_DE_HABILIDAD_{m.assessment_number}": m for m in mapping
        }

        student_plans: dict[tuple[str, str], ExamenPlan] = {}

        for assessment_name, df in download_result.items():
            map_row = mapping_by_assessment_name.get(assessment_name)
            if map_row is None:
                continue

            bank_df = pd.read_excel(
                BytesIO(self._read_bank_bytes(map_row.bank_path.name, local_path=map_row.bank_path))
            )
            bank_df.columns = [_normalize_text(c).lower() for c in bank_df.columns]
            required_cols = {"pregunta", "alternativa", "unidad"}
            if not required_cols.issubset(set(bank_df.columns)):
                missing = required_cols - set(bank_df.columns)
                raise ValueError(f"Bank {map_row.bank_path.name} missing columns: {sorted(missing)}")

            for _, student_row in df.iterrows():
                email = _normalize_text(student_row.get("email") or student_row.get("username") or "")
                student_id = _normalize_text(student_row.get("user_id") or email)
                if not email:
                    continue

                plan_key = (map_row.assessment_type, email)
                if plan_key not in student_plans:
                    student_plans[plan_key] = ExamenPlan(
                        assessment_type=map_row.assessment_type,
                        student_id=student_id,
                        email=email,
                        assessment_name=map_row.assessment_name,
                    )
                plan = student_plans[plan_key]

                for _, question in bank_df.iterrows():
                    unit_name = _display_text(question["unidad"])
                    question_label = _display_text(question["pregunta"])
                    correct_answer = _normalize_answer(question["alternativa"])
                    user_answer = _normalize_answer(_lookup_answer_column(student_row, question_label))
                    is_correct = user_answer == correct_answer

                    if unit_name not in plan.units:
                        plan.units[unit_name] = UnitStats(name=unit_name)
                        plan.unit_order.append(unit_name)
                    unit = plan.units[unit_name]
                    unit.total += 1
                    unit.correct += 1 if is_correct else 0

        return student_plans

    def render(self, analysis_result: dict[tuple[str, str], ExamenPlan]) -> Path:
        output_dir = self.data_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        cover_html = (self.templates_path / "cover.html").read_text(encoding="utf-8")
        body_template = load_body_template(REPORT_TYPE)

        # Discover which placeholder keys actually appear in body.html
        schema = load_report_placeholder_schema(REPORT_TYPE)
        used_in_body = discover_placeholders_in_html(body_template)
        computed_allowed = set(schema["computed"])
        static_allowed = set(schema["static"])

        for (_, email), plan in analysis_result.items():
            ordered_units = [plan.units[u] for u in plan.unit_order if u in plan.units]
            if not ordered_units:
                continue

            unit_rows = [
                {
                    "unidad": u.name,
                    "estado": _assign_estado_recomendacion(u.percent)[0],
                    "recomendacion": _assign_estado_recomendacion(u.percent)[1],
                }
                for u in ordered_units
            ]

            # Build only the intersection of (schema keys) and (keys used in body.html)
            computed_values: dict[str, Any] = {}
            static_values: dict[str, Any] = {}

            candidate_computed = {
                "student_name": _normalize_text(email),
                "course_name": _strip_data_suffix(plan.assessment_name or plan.assessment_type),
                "generated_at": datetime.now().strftime("%Y-%m-%d"),
                "period_label": "",
            }
            for k, v in candidate_computed.items():
                if k in used_in_body and k in computed_allowed:
                    computed_values[k] = v

            candidate_static: dict[str, Any] = {}
            for k in static_allowed:
                candidate_static[k] = ""
            for k, v in candidate_static.items():
                if k in used_in_body and k in static_allowed:
                    static_values[k] = v

            rendered_body = render_with_placeholders(
                report_type=REPORT_TYPE,
                body_html=body_template,
                computed_values=computed_values,
                static_values=static_values,
            )

            rendered_body = insert_dynamic_tables(
                REPORT_TYPE,
                rendered_body,
                {"unit_status_rows": unit_rows},
            )

            final_html = _compose_cover_plus_body_html(cover_html, rendered_body)

            pdf_bytes = HTML(string=final_html, base_url=str(Path.cwd())).write_pdf()
            assessment_label = plan.assessment_name or plan.assessment_type
            pdf_path = output_dir / (
                f"informe_{_safe_filename_component(REPORT_TYPE)}"
                f"_{_safe_filename_component(assessment_label)}"
                f"_{_safe_filename_component(email)}.pdf"
            )
            pdf_path.write_bytes(pdf_bytes)

        return output_dir
