# Copy of the existing analisis.py file for Cloud Function
# This file will be copied from the main directory during deployment

import os
import json
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import argparse
from typing import Dict, List, Any
import yaml
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def load_course_config(config_path: str = "cursos.yml"):
    """Load course configuration from YAML file"""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def load_processed_data(course_id: str) -> Dict[str, pd.DataFrame]:
    """Load processed CSV data for a course"""
    processed_dir = Path(f"data/processed/{course_id}")
    
    data = {}
    for file_type in ['assessments', 'grades', 'users']:
        file_path = processed_dir / f"{file_type}.csv"
        if file_path.exists():
            data[file_type] = pd.read_csv(file_path)
        else:
            print(f"Warning: {file_path} not found")
            data[file_type] = pd.DataFrame()
    
    return data

def calculate_kpis(data: Dict[str, pd.DataFrame], course_config: Dict[str, Any]) -> Dict[str, Any]:
    """Calculate KPIs for the course"""
    kpis = {}
    
    # Get KPI list from config
    kpi_list = course_config.get('kpis', [])
    
    if 'attendance_rate' in kpi_list:
        kpis['attendance_rate'] = calculate_attendance_rate(data)
    
    if 'average_grade' in kpi_list:
        kpis['average_grade'] = calculate_average_grade(data)
    
    if 'completion_rate' in kpi_list:
        kpis['completion_rate'] = calculate_completion_rate(data)
    
    if 'response_rate' in kpi_list:
        kpis['response_rate'] = calculate_response_rate(data)
    
    return kpis

def calculate_attendance_rate(data: Dict[str, pd.DataFrame]) -> float:
    """Calculate attendance rate"""
    if 'users' in data and not data['users'].empty:
        total_users = len(data['users'])
        active_users = len(data['users'][data['users']['last_login_at'].notna()])
        return (active_users / total_users) * 100 if total_users > 0 else 0
    return 0

def calculate_average_grade(data: Dict[str, pd.DataFrame]) -> float:
    """Calculate average grade"""
    if 'grades' in data and not data['grades'].empty:
        grades = data['grades']['grade'].dropna()
        return grades.mean() if len(grades) > 0 else 0
    return 0

def calculate_completion_rate(data: Dict[str, pd.DataFrame]) -> float:
    """Calculate completion rate"""
    if 'assessments' in data and not data['assessments'].empty:
        total_assessments = len(data['assessments'])
        completed_assessments = len(data['assessments'][data['assessments']['completed_at'].notna()])
        return (completed_assessments / total_assessments) * 100 if total_assessments > 0 else 0
    return 0

def calculate_response_rate(data: Dict[str, pd.DataFrame]) -> float:
    """Calculate response rate"""
    if 'grades' in data and not data['grades'].empty:
        total_grades = len(data['grades'])
        responded_grades = len(data['grades'][data['grades']['grade'].notna()])
        return (responded_grades / total_grades) * 100 if total_grades > 0 else 0
    return 0

def save_kpis_to_csv(kpis: Dict[str, Any], course_id: str):
    """Save KPIs to CSV file"""
    kpi_dir = Path(f"data/metrics/kpi")
    kpi_dir.mkdir(parents=True, exist_ok=True)
    
    kpi_df = pd.DataFrame([kpis])
    kpi_df.to_csv(kpi_dir / f"{course_id}.csv", index=False)

def generate_pdf_report(course_id: str, course_config: Dict[str, Any], kpis: Dict[str, Any], data: Dict[str, pd.DataFrame]):
    """Generate PDF report"""
    reports_dir = Path(f"data/reports/{course_id}")
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, f'Reporte de Análisis - {course_config.get("name", course_id)}', ln=True, align='C')
    pdf.ln(10)
    
    # Date
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'Fecha de generación: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', ln=True)
    pdf.ln(10)
    
    # KPIs
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Indicadores Clave de Rendimiento (KPIs)', ln=True)
    pdf.ln(5)
    
    pdf.set_font('Arial', '', 12)
    for kpi_name, kpi_value in kpis.items():
        if isinstance(kpi_value, float):
            pdf.cell(0, 10, f'{kpi_name.replace("_", " ").title()}: {kpi_value:.2f}%', ln=True)
        else:
            pdf.cell(0, 10, f'{kpi_name.replace("_", " ").title()}: {kpi_value}', ln=True)
    
    pdf.ln(10)
    
    # Data summary
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Resumen de Datos', ln=True)
    pdf.ln(5)
    
    pdf.set_font('Arial', '', 12)
    for data_type, df in data.items():
        if not df.empty:
            pdf.cell(0, 10, f'{data_type.title()}: {len(df)} registros', ln=True)
    
    # Save PDF
    pdf_path = reports_dir / f"reporte_{course_id}.pdf"
    pdf.output(str(pdf_path))
    return pdf_path

def generate_excel_report(course_id: str, course_config: Dict[str, Any], kpis: Dict[str, Any], data: Dict[str, pd.DataFrame]):
    """Generate Excel report"""
    reports_dir = Path(f"data/reports/{course_id}")
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # KPI Summary sheet
    ws_kpis = wb.active
    ws_kpis.title = "KPIs"
    
    # Headers
    headers = ['KPI', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws_kpis.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # KPI data
    for row, (kpi_name, kpi_value) in enumerate(kpis.items(), 2):
        ws_kpis.cell(row=row, column=1, value=kpi_name.replace('_', ' ').title())
        if isinstance(kpi_value, float):
            ws_kpis.cell(row=row, column=2, value=f"{kpi_value:.2f}%")
        else:
            ws_kpis.cell(row=row, column=2, value=kpi_value)
    
    # Data sheets
    for data_type, df in data.items():
        if not df.empty:
            ws = wb.create_sheet(title=data_type.title())
            
            # Add data
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save Excel
    excel_path = reports_dir / f"reporte_{course_id}.xlsx"
    wb.save(str(excel_path))
    return excel_path

def generate_listas_excel(course_id: str, data: Dict[str, pd.DataFrame]):
    """Generate Excel file with lists for each data type"""
    reports_dir = Path(f"data/reports/{course_id}")
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets for each data type
    for data_type, df in data.items():
        if not df.empty:
            ws = wb.create_sheet(title=data_type.title())
            
            # Add data
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save Excel
    excel_path = reports_dir / f"reporte_listas_{course_id}.xlsx"
    wb.save(str(excel_path))
    return excel_path

def run_analysis_pipeline(course_id: str):
    """Run the complete analysis pipeline for a course"""
    try:
        # Load configuration
        config = load_course_config()
        courses = config.get('courses', {})
        
        if course_id not in courses:
            raise ValueError(f"Course {course_id} not found in configuration")
        
        course_config = courses[course_id]
        
        # Load processed data
        data = load_processed_data(course_id)
        
        # Calculate KPIs
        kpis = calculate_kpis(data, course_config)
        
        # Save KPIs
        save_kpis_to_csv(kpis, course_id)
        
        # Generate reports
        pdf_path = generate_pdf_report(course_id, course_config, kpis, data)
        excel_path = generate_excel_report(course_id, course_config, kpis, data)
        listas_path = generate_listas_excel(course_id, data)
        
        print(f"✅ Analysis completed for {course_id}")
        print(f"  - PDF: {pdf_path}")
        print(f"  - Excel: {excel_path}")
        print(f"  - Listas: {listas_path}")
        
    except Exception as e:
        print(f"❌ Analysis failed for {course_id}: {str(e)}")
        raise

def main():
    parser = argparse.ArgumentParser(description='Analyze course data and generate reports')
    parser.add_argument('--course', '-c', required=True, help='Course ID to analyze')
    
    args = parser.parse_args()
    run_analysis_pipeline(args.course)

if __name__ == "__main__":
    main() 