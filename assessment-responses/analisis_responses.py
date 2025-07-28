import os
import argparse
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
import subprocess
from storage import StorageClient

load_dotenv()

def load_correct_answers(csv_path):
    # Try to load <assessment>_questions.csv from the questions directory
    # The CSV path is in processed/<course>/, so we need to get the course name
    course = csv_path.parent.name
    base = csv_path.stem.split('_')[0]
    questions_dir = csv_path.parent.parent.parent / "questions" / course
    raw_question_file = questions_dir / f"{base}.csv"
    questions_file = questions_dir / f"{base}_questions.csv"
    # Check for raw question file
    if not questions_file.exists():
        if raw_question_file.exists():
            print(f"Found raw question file: {raw_question_file}. Converting to {questions_file}...")
            # Run convert_questions_file.py
            result = subprocess.run([
                "python", "convert_questions_file.py",
                "--input", str(raw_question_file),
                "--output", str(questions_file)
            ])
            if result.returncode != 0 or not questions_file.exists():
                print(f"Failed to convert {raw_question_file} to {questions_file}. Please check the input file.")
                return {}
        else:
            print(f"Missing question file: {raw_question_file}. Please upload it to {questions_dir} and rerun.")
            exit(1)
    try:
        storage = StorageClient()
        df = storage.read_csv(questions_file)
        # Expect columns: question, correct_answer
        return dict(zip(df['question'], df['correct_answer']))
    except Exception as e:
        print(f"Could not load correct answers from {questions_file}: {e}")
        return {}

def generate_question_alt_stats(df, question_cols, correct_answers, alt_map, all_alts):
    stats = []
    for q in question_cols:
        answers = df[q].dropna().astype(str)
        total = len(answers)
        correct_answer = correct_answers.get(q, '')
        alternatives = alt_map.get(q, [])
        # Use all_alts for consistent columns
        alt_counts = {alt: 0 for alt in all_alts}
        for ans in answers:
            if ans in alt_counts:
                alt_counts[ans] += 1
        alt_percents = {alt: (alt_counts[alt] / total * 100 if total > 0 else 0) for alt in all_alts}
        # Find the alternative with the highest percentage
        most_selected = max(alt_percents, key=alt_percents.get) if alt_percents else ''
        stats.append({
            'question': q,
            'cantidad': total,
            'respuesta_correcta': correct_answer,
            'alternatives': all_alts,
            'alt_percents': alt_percents,
            'most_selected': most_selected
        })
    return stats

def generate_pdf_report(df_stats, output_path, top_percent, assessment_name):
    from fpdf import FPDF
    # Build dynamic headers
    if not df_stats:
        return
    # Collect all unique alternatives across all questions, sorted for consistency
    all_alts = df_stats[0]['alternatives'] if df_stats else []
    headers = ["Pregunta", "Cantidad", "Respuesta Correcta"] + all_alts
    # Prepare table data
    table_data = []
    for s in df_stats:
        row = [
            str(s['question'])[:100],
            str(s['cantidad']),
            str(s['respuesta_correcta'])[:100]
        ]
        for alt in all_alts:
            row.append(f"{s['alt_percents'].get(alt, 0):.1f}%")
        table_data.append(row)
    # Calculate correct answer percentage for sorting
    correct_col_idx = 2 + all_alts.index(df_stats[0]['respuesta_correcta']) if df_stats and df_stats[0]['respuesta_correcta'] in all_alts else None
    correct_percents = []
    for s in df_stats:
        ca = s['respuesta_correcta']
        correct_percents.append(s['alt_percents'].get(ca, 0))
    # Sort by correct answer percentage ascending
    sorted_indices = sorted(range(len(df_stats)), key=lambda i: correct_percents[i])
    table_data = [table_data[i] for i in sorted_indices]
    df_stats = [df_stats[i] for i in sorted_indices]
    # Calculate column widths dynamically
    min_width, max_width = 18, 60
    col_widths = []
    for col_idx in range(len(headers)):
        max_len = len(str(headers[col_idx]))
        for row in table_data:
            max_len = max(max_len, len(str(row[col_idx])))
        # Approximate width: 2.5 per char, clamp to min/max
        width = min(max_width, max(min_width, int(max_len * 2.5)))
        col_widths.append(width)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, assessment_name, ln=1)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f'Top {top_percent}% Preguntas Más Falladas', ln=1)
    pdf.set_font('Arial', 'B', 9)
    # Table header
    x_start = pdf.get_x()
    y_start = pdf.get_y()
    max_height = 0
    for i, header in enumerate(headers):
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.multi_cell(col_widths[i], 5, header, border=1, align='C')
        max_height = max(max_height, pdf.get_y() - y)
        pdf.set_xy(x + col_widths[i], y)
    pdf.set_xy(x_start, y_start + max_height)
    pdf.set_font('Arial', '', 8)
    # Only top x% questions
    n_top = max(1, round(len(table_data) * top_percent / 100))
    for idx, row in enumerate(table_data[:n_top]):
        s = df_stats[idx]
        highlight = s['respuesta_correcta'] and s['respuesta_correcta'] != s['most_selected']
        if highlight:
            pdf.set_fill_color(255, 200, 200)  # soft red
        for i, value in enumerate(row):
            pdf.cell(col_widths[i], 8, value, border=1, align='C', fill=highlight)
        pdf.ln()
        if highlight:
            pdf.set_fill_color(255, 255, 255)
    pdf.output(output_path)

def generate_xlsx_report(df_stats, output_path, assessment_name):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    # Build dynamic headers
    if not df_stats:
        return
    all_alts = df_stats[0]['alternatives'] if df_stats else []
    headers = ["Pregunta", "Cantidad", "Respuesta Correcta"] + all_alts
    rows = []
    correct_percents = []
    for s in df_stats:
        row = [s['question'], s['cantidad'], s['respuesta_correcta']]
        for alt in all_alts:
            row.append(s['alt_percents'].get(alt, 0))
        rows.append(row)
        ca = s['respuesta_correcta']
        correct_percents.append(s['alt_percents'].get(ca, 0))
    # Sort by correct answer percentage ascending
    sorted_indices = sorted(range(len(df_stats)), key=lambda i: correct_percents[i])
    rows = [rows[i] for i in sorted_indices]
    df_stats = [df_stats[i] for i in sorted_indices]
    df = pd.DataFrame(rows, columns=headers)
    df.to_excel(output_path, index=False)
    # Highlight entire row if correct answer is not the most selected
    wb = load_workbook(output_path)
    ws = wb.active
    correct_fill = PatternFill(start_color="FFF0C0C0", end_color="FFF0C0C0", fill_type="solid")
    for i, s in enumerate(df_stats, start=2):
        if s['respuesta_correcta'] and s['respuesta_correcta'] != s['most_selected']:
            for col in range(1, len(headers)+1):
                ws.cell(row=i, column=col).fill = correct_fill
    wb.save(output_path)

def process_csv(csv_path, top_percent):
    storage = StorageClient()
    # Use storage to read CSV
    df = storage.read_csv(str(csv_path), sep=';')
    question_cols = [col for col in df.columns if str(col).lower().startswith('pregunta ')]
    if not question_cols:
        print(f"No question columns found in {csv_path}")
        return
    correct_answers = load_correct_answers(csv_path)
    # Build alternative map from correct answers file
    course = csv_path.parent.name
    base = csv_path.stem.split('_')[0]
    questions_dir = csv_path.parent.parent.parent / "questions" / course
    questions_file = questions_dir / f"{base}_questions.csv"
    alt_map = {}
    all_alts = set()
    if questions_file.exists():
        qdf = pd.read_csv(questions_file)
        for _, row in qdf.iterrows():
            q = row['question']
            alts = [row[c] for c in qdf.columns if c.lower().startswith('answer') and pd.notnull(row[c])]
            alt_map[q] = alts
            all_alts.update(alts)
    # Also scan all student answers for any alternatives not in the questions file
    for q in question_cols:
        answers = df[q].dropna().astype(str)
        all_alts.update(answers.unique())
    all_alts = sorted(all_alts)
    df_stats = generate_question_alt_stats(df, question_cols, correct_answers, alt_map, all_alts)
    # Save reports in the reports/<course>/ directory
    course = csv_path.parent.name
    reports_dir = csv_path.parent.parent.parent / "reports" / course
    reports_dir.mkdir(parents=True, exist_ok=True)
    assessment_name = base
    # XLSX report
    import tempfile
    xlsx_filename = f"{csv_path.stem}_report.xlsx"
    pdf_filename = f"{csv_path.stem}_top_{top_percent}pct.pdf"
    if storage.backend == 'local':
        xlsx_path = reports_dir / xlsx_filename
        generate_xlsx_report(df_stats, xlsx_path, assessment_name)
        pdf_path = reports_dir / pdf_filename
        generate_pdf_report(df_stats, pdf_path, top_percent, assessment_name)
    else:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_xlsx:
            generate_xlsx_report(df_stats, tmp_xlsx.name, assessment_name)
            tmp_xlsx.flush()
            with open(tmp_xlsx.name, 'rb') as f:
                storage.write_bytes(f"data/responses/reports/{course}/{xlsx_filename}", f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
            generate_pdf_report(df_stats, tmp_pdf.name, top_percent, assessment_name)
            tmp_pdf.flush()
            with open(tmp_pdf.name, 'rb') as f:
                storage.write_bytes(f"data/responses/reports/{course}/{pdf_filename}", f.read(), content_type='application/pdf')
    print(f"Reports generated for {csv_path}")

def parse_arguments():
    parser = argparse.ArgumentParser(description='Generate analysis and reports for assessment responses.')
    parser.add_argument('--course', '-c', required=True, help='Course ID to process')
    parser.add_argument('--assessment', '-a', action='append', help='Assessment name to analyze (can be used multiple times)')
    parser.add_argument('--all', action='store_true', help='Analyze all assessments for the course')
    return parser.parse_args()

def main():
    args = parse_arguments()
    try:
        top_percent = int(os.getenv('REPORT_TOP_PERCENT', 20))
        if not (1 <= top_percent <= 100):
            print('REPORT_TOP_PERCENT out of range, using 20')
            top_percent = 20
    except Exception:
        top_percent = 20
    processed_dir = Path('data/responses/processed') / args.course
    if not processed_dir.exists():
        print(f"Processed directory not found: {processed_dir}")
        return
    if args.all:
        csv_files = list(processed_dir.glob('*.csv'))
    elif args.assessment:
        csv_files = [processed_dir / f"{name.replace('/', '_')}.csv" for name in args.assessment]
        missing = [str(f) for f in csv_files if not f.exists()]
        if missing:
            print(f"Warning: The following assessment CSVs were not found: {missing}")
        csv_files = [f for f in csv_files if f.exists()]
    else:
        print("Error: You must specify either --all or at least one --assessment.")
        return
    for csv_file in csv_files:
        process_csv(csv_file, top_percent)

if __name__ == '__main__':
    main() 